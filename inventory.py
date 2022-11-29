# imports
import datetime as dt
import sqlite3
import sys
import tkinter as tk
import warnings
from tkinter import *
from tkinter import ttk
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.simplefilter('ignore')

# database path
db = '//cncd-dc1/Laboratory/Chemistry/inv_database/biochem.sqlite'

# Creating folder to save downloads
if not os.path.isdir("saves"):
    # if not present then create it.
    os.mkdir("saves")

# Creating a main window
window = tk.Tk()

# Modifying window theme
window.title("Inventory Management System (by: Shahzaib Raza)")
window.geometry("1000x720")
window.configure(bg='#323232')

# Adding title to the main window
in_title= tk.Label(
    text="Inventory Management System",
    fg='white',
    bg="#323232",
    font=("Arial", 25)
    )
in_title.pack()

# *******************************************************************************************************

# Left Frame
left_frame = tk.Frame(
    window,
    bg='#323232',
    highlightbackground ="grey",
    highlightthickness=2,
    )

# *******************************************************************************************************

# Function to get table names of database
def get_cols():
    global db
    conn = sqlite3.connect(db)

    q = """
        SELECT name FROM sqlite_schema
        WHERE type='table'
        ORDER BY name;
    """

    cur = conn.cursor()

    r = cur.execute(q)

    cols = [i[0] for i in r.fetchall()]

    conn.close()

    return cols

# *******************************************************************************************************

# table input
ti_1 = tk.Frame(
    left_frame,
    bg='#323232'
)

tilab1 = tk.Label(
    ti_1,
    text="Table:",
    fg='white',
    bg="#323232",
)
tilab1.pack(side=tk.TOP)

# Dropdown menu options
opt = get_cols()

# datatype of menu text
click = StringVar()

click.set(opt[0])

dr = OptionMenu(
    ti_1,
    click,
    *opt,
    )
dr.pack(padx=5, side=tk.LEFT)

ti_1.pack(side=tk.TOP)

# *******************************************************************************************************

separator_0 = ttk.Separator(left_frame, orient='horizontal')
separator_0.pack(fill='x', pady=40)

# *******************************************************************************************************

# Second input
lf_2 = tk.Frame(
    left_frame,
    bg='#323232'
)

lab2 = tk.Label(
    lf_2,
    text="Brand Name:",
    fg='white',
    bg="#323232",
)
lab2.pack()

inp2 = tk.Entry(
    lf_2,
    width=30
    )
inp2.pack(padx=5)

lf_2.pack(side=tk.TOP)

# *******************************************************************************************************

# Third input
lf_3 = tk.Frame(
    left_frame,
    bg='#323232'
)

lab3 = tk.Label(
    lf_3,
    text="Quantity In:",
    fg='white',
    bg="#323232",
)
lab3.pack()

inp3 = tk.Entry(
    lf_3,
    width=30,
    )
inp3.pack(padx=5)

lf_3.pack(side=tk.TOP)

# *******************************************************************************************************

# Fourth input
lf_4 = tk.Frame(
    left_frame,
    bg='#323232'
)

lab4 = tk.Label(
    lf_4,
    text="Quantity Out:",
    fg='white',
    bg="#323232",
)
lab4.pack()

inp4 = tk.Entry(
    lf_4,
    width=30,
    )
inp4.pack(padx=5)

lf_4.pack(side=tk.TOP)

# *******************************************************************************************************

# Fifth input
lf_5 = tk.Frame(
    left_frame,
    bg='#323232'
)

lab5 = tk.Label(
    lf_5,
    text="Balance:",
    fg='white',
    bg="#323232",
)
lab5.pack()

inp5 = tk.Entry(
    lf_5,
    width=30,
    )
inp5.pack()

lf_5.pack(side=tk.TOP)

# *******************************************************************************************************

# Sixth input
lf_6 = tk.Frame(
    left_frame,
    bg='#323232'
)

lab6 = tk.Label(
    lf_6,
    text="Name of staff:",
    fg='white',
    bg="#323232",
)
lab6.pack()

inp6 = tk.Entry(
    lf_6,
    width=30,
    )
inp6.pack(padx=5)

lf_6.pack(side=tk.TOP)

# *******************************************************************************************************

# Seventh input
lf_7 = tk.Frame(
    left_frame,
    bg='#323232'
)

lab7 = tk.Label(
    lf_7,
    text="Remarks:",
    fg='white',
    bg="#323232",
)
lab7.pack()

inp7 = tk.Entry(
    lf_7,
    width=30,
    )
inp7.pack(padx=5)

lf_7.pack(side=tk.TOP)

# *******************************************************************************************************

message = tk.Label(
    left_frame,
    font=('Arial', 8,),
    fg='white',
    bg='#323232',
)

message.pack(side=tk.TOP, padx=5, fill=tk.BOTH, pady=5)

# *******************************************************************************************************

# Function to get table data
def get_table(name: str):

    global db

    conn = sqlite3.connect(db)

    q = f"""
        SELECT * FROM '{name}';
    """

    df = pd.read_sql(sql=q, con=conn)

    df['id'] = df['id'].astype('int')

    conn.close()

    return df

# *******************************************************************************************************

def insert_int():
    global message
    global click
    global inp2
    global inp3
    global inp4
    global inp5
    global inp6
    global inp7
    global db

    vals = [inp2, inp3, inp4, inp5]
    
    tb = click.get()
    
    dta = get_table(tb)

    new_i = int(dta['id'].to_list()[-1]) + 1
    
    go = True
    for v in vals:
        if v.get() == "":
            go = False
            message.config(text="Entry can not be blank.")
            break
    
    inp1 = dt.datetime.now().date().strftime("%d/%m/%Y")

    if go == True:

        i6, i7 = inp6.get(), inp7.get()
        
        nas = 0

        if i6 == '' and i7 == '':
            qry = f"""
                INSERT INTO '{tb}'
                VALUES ({new_i}, '{inp1}', '{inp2.get()}', {inp3.get()}, {inp4.get()}, {inp5.get()}, ?, ?);
            """
            nas = 2
        elif i6 == '':
            qry = f"""
                INSERT INTO '{tb}'
                VALUES ({new_i}, '{inp1}', '{inp2.get()}', {inp3.get()}, {inp4.get()}, {inp5.get()}, ?, '{i7}');
            """
            nas = 1
        
        elif i7 == '':
            qry = f"""
                INSERT INTO '{tb}'
                VALUES ({new_i}, '{inp1}', '{inp2.get()}', {inp3.get()}, {inp4.get()}, {inp5.get()}, '{i6}', ?);
            """
            nas = 1
        
        else:
            qry = f"""
                INSERT INTO '{tb}'
                VALUES ({new_i}, '{inp1}', '{inp2.get()}', {inp3.get()}, {inp4.get()}, {inp5.get()}, '{i6}', '{i7}');
            """

        message.config(text="Successfully entered the record.")
        
        conn = sqlite3.connect(db)

        cur = conn.cursor()

        cur.execute(qry, tuple(None for i in range(nas)))
        
        conn.commit()

        conn.close()

# *******************************************************************************************************

def defaults():
    global db
    tabls = get_cols()

    conn = sqlite3.connect(db)

    cur = conn.cursor()
    
    inp1 = dt.datetime.now().date().strftime("%d/%m/%Y")

    for tab in tabls:
        
        dta = get_table(tab)

        if dta['Date'].to_list()[-1] != inp1:

            new_i = int(dta['id'].to_list()[-1]) + 1

            bal = dta['Balance'].to_list()[-1]

            qry = f"""
                INSERT INTO '{tab}' VALUES ({new_i}, '{inp1}', '{tab}', 0, 0, {bal}, ?, ?)
            """
            try:
                cur.execute(qry, tuple(None for i in range(2)))
            except:
                print("********************************")
                print(f"Unable to insert in table: {tab}")
                print("********************************")
    
    conn.commit()

    conn.close()
    
    message.config(text="Successfully entered the default records.")

    ins_button.config(bg="#4F7942")

# *******************************************************************************************************

ins_button  = tk.Button(
    left_frame,
    text='Insert Defaults',
    fg='white',
    bg="#8B0000",
    width=11,
    height=1,
    command=defaults
)

ins_button.pack(side=tk.TOP)

# *******************************************************************************************************

inp_button  = tk.Button(
    left_frame,
    text='Insert',
    fg='white',
    bg="#4F7942",
    width=8,
    height=1,
    command=insert_int
)

inp_button.pack(side=tk.TOP, pady=5)

# *******************************************************************************************************

separator = ttk.Separator(left_frame, orient='horizontal')
separator.pack(fill='x', pady=20)

# *******************************************************************************************************

# Eighth input
lf_8 = tk.Frame(
    left_frame,
    bg='#323232'
)

lab8 = tk.Label(
    lf_8,
    text="ID:",
    fg='white',
    bg="#323232",
)
lab8.pack()

inp8 = tk.Entry(
    lf_8,
    width=30,
    )
inp8.pack(padx=5)

lf_8.pack(side=tk.TOP, pady=5)

# *******************************************************************************************************

def del_id():
    global inp8
    global click
    global db

    try:
        id_ = int(inp8.get())
        tab = click.get()
        dq = f"""
             DELETE FROM '{tab}' WHERE id = {id_};
        """

        conn = sqlite3.connect(db)

        cur = conn.cursor()

        cur.execute(dq)

        conn.commit()

        conn.close()

        message.config(text="Successfully deleted the record.")

    except:
        pass


# *******************************************************************************************************

del_button  = tk.Button(
    left_frame,
    text='Delete',
    fg='white',
    bg="#8B0000",
    width=8,
    height=1,
    command=del_id
)

del_button.pack(side=tk.TOP, pady=5)

# *******************************************************************************************************

left_frame.pack(side=tk.LEFT, anchor=tk.NW, padx=5, fill=tk.BOTH, pady=30)


# *******************************************************************************************************

top_frame = tk.Frame(
    window,
    bg='#323232',
    highlightbackground ="grey",
    highlightthickness=2,
    )


# *******************************************************************************************************

# First input
tf_1 = tk.Frame(
    top_frame,
    bg='#323232'
)

tlab1 = tk.Label(
    tf_1,
    text="Table:",
    fg='white',
    bg="#323232",
)
tlab1.pack(side=tk.LEFT)

# Dropdown menu options
options = get_cols()

# datatype of menu text
clicked = StringVar()

clicked.set(options[0])

drop = OptionMenu(
    tf_1,
    clicked,
    *options,
    )
drop.pack(padx=5, side=tk.LEFT)

tf_1.pack(side=tk.LEFT)

# *******************************************************************************************************

status = tk.Label(
    top_frame,
    font=('Arial', 12, 'bold'),
    fg='white',
    text="Clear",
    bg='#4F7942',
)

# *******************************************************************************************************

s_frame = tk.Frame(
    window,
    bg='#323232',
    highlightbackground ="grey",
    highlightthickness=2,
    height=1200
    )

# *******************************************************************************************************

tx = tk.Text(
    s_frame,
    width=500,
    height=1190,
    fg='white',
    bg='#323232'
    )

# *******************************************************************************************************

class PrintToTXT(object): 
    def write(self, s): 
        tx.insert(END, s)
    def flush(self, s):
        tx.delete("1.0","end")

# *******************************************************************************************************

# Function for printing data on screen
def show_inv():

    global clicked
    tab = clicked.get()
    
    tx.delete("1.0","end")

    dta = get_table(tab)

    dta['Qua_In'] = dta['Qua_In'].replace(np.nan, 0).astype('int')
    dta['Qua_Out'] = dta['Qua_Out'].replace(np.nan, 0).astype('int')
    dta['Balance'] = dta['Balance'].replace(np.nan, 0).astype('int')
    
    s1 = dta['Qua_In'].sum()
    s2 = dta['Qua_Out'].sum()
    
    if (s1-s2) < 5:
        status.config(
            text="Need to order",
            bg='#8B0000',
        )
        tx.config(bg='#8B0000')
    else:
        status.config(
            text="Clear",
            bg='#4F7942',
        )
        tx.config(bg='#4F7942')

    sd = dta.to_string()

    sys.stdout = PrintToTXT()

    print(sd)

    return 1

# *******************************************************************************************************

get_button  = tk.Button(
    top_frame,
    text='Show Inventory',
    fg='white',
    bg="#4682B4",
    height=1,
    command=show_inv
)

get_button.pack(side=tk.LEFT, padx=5)

# *******************************************************************************************************

def down_inv():
    
    global clicked
    tab = clicked.get()
    
    show_inv()

    mnth = dt.datetime.now().date().month

    mnth_str = {1: "JAN", 2: "FEB", 3: "MARCH", 4:"APR", 5:"MAY", 6:"JUN",
        7: "JUL", 8: "AUG", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC"
    }

    sm = mnth_str[mnth]

    fn = tab + "_" + sm + ".xlsx"

    tbl = get_table(tab)

    tbl['month'] = pd.DatetimeIndex(tbl['Date']).month

    tbl = tbl[tbl['month'] == mnth]

    tbl.drop(columns=['month'], inplace=True)

    # s1 = tbl['Qua_In'].sum()
    # s2 = tbl['Qua_Out'].sum()
     
    writer = pd.ExcelWriter(f'saves/{fn}')

    tbl.to_excel(writer, index=False, startrow=2)

    writer.save()

    writer.close()

    book = load_workbook(f'saves/{fn}')
    sheet = book['Sheet1']
    sheet.cell(row=1, column=1).value = "Status:"
    sheet.cell(row=1, column=2).value = status.cget("text")
    
    book.save(f'saves/{fn}')

    book.close()

down_button = tk.Button(
    top_frame,
    text='Download this month data',
    fg='white',
    bg="#008080",
    height=1,
    command=down_inv
)

down_button.pack(side=tk.LEFT, padx=8)

# *******************************************************************************************************

status.pack(side=tk.RIGHT, padx=5)

# *******************************************************************************************************

top_frame.pack(side=tk.TOP, anchor=tk.N, padx=5, fill=tk.BOTH, pady=5)

# *******************************************************************************************************

s_frame.pack(side=tk.TOP, anchor=tk.S, padx=5, fill=tk.BOTH, pady=5)

# *******************************************************************************************************

tx.pack()

# *******************************************************************************************************

# Running the window
window.mainloop()